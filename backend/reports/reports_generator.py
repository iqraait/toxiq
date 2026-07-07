from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.utils import timezone
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

def style_header_row(ws, columns):
    ws.append(columns)
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1A365D', end_color='1A365D', fill_type='solid')
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='medium', color='1E293B')
    )
    
    for col_idx in range(1, len(columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    ws.row_dimensions[1].height = 28

def auto_fit_columns(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

def generate_registrations_excel(registrations):
    wb = Workbook()
    ws = wb.active
    ws.title = "Registrations"
    
    from registration.models import RegistrationField, RegistrationForm
    active_form = RegistrationForm.objects.filter(is_active=True).first()
    form_fields = list(active_form.fields.all().order_by('order', 'id')) if active_form else []
    
    headers = [
        "Registration ID", 
        "Full Name", 
        "Email", 
        "Phone", 
        "Designation",
        "Institute / Hospital",
        "Specialty / Department of Practice",
        "Registration Category",
        "Medical Council Name",
        "Registration No:",
        "Food Preference",
        "Paid Amount",
        "Payment Status", 
        "Transaction ID", 
        "Created Date"
    ]
    
    style_header_row(ws, headers)
    
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    
    for row_idx, reg in enumerate(registrations, start=2):
        payment = reg.payments.filter(payment_status='SUCCESS').first()
        if not payment:
            payment = reg.payments.first()
            
        status = payment.payment_status if payment else "NO_PAYMENT"
        txnid = payment.transaction_id if payment else "N/A"
        amount = f"{payment.currency} {payment.amount:.2f}" if payment else "N/A"
        
        reg_date = reg.created_at.astimezone(timezone.get_current_timezone()).strftime("%Y-%m-%d %H:%M")
        
        def get_val_dynamic(keywords):
            field_data = reg.field_data or {}
            values = list(field_data.values())

            prefix_options = ['mr.', 'ms.', 'mrs.', 'dr.', 'mr', 'ms', 'mrs', 'dr']
            food_options = ['veg', 'non-veg', 'veg.', 'non-veg.']
            specialty_keywords = ['emergency', 'pharmacy', 'critical', 'care', 'medicine', 'pediatrics', 'forensic', 'practitioner', 'specialty', 'department']
            category_keywords = ['speaker', 'consultant', 'resident', 'student', 'intern', 'nurse', 'pharmacist', 'paramedic', 'delegate', 'category']
            designation_keywords = ['doctor', 'consultant', 'physician', 'resident', 'student', 'intern', 'nurse', 'pharmacist', 'paramedic', 'professor', 'lecturer', 'director', 'manager', 'specialist', 'practitioner', 'hod', 'registrar', 'fellow', 'designation']
            council_keywords = ['council', 'medical', 'imc', 'kcl', 'mc', 'karnataka', 'delhi', 'maharashtra', 'registration council', 'state']

            email_val = ''
            phone_val = ''
            prefix_val = ''
            food_val = ''
            specialty_val = ''
            category_val = ''
            reg_no_val = ''
            designation_val = ''
            council_val = ''
            name_val = ''
            institute_val = ''

            import re

            for raw_val in values:
                if raw_val is None or raw_val == '':
                    continue

                if isinstance(raw_val, list):
                    joined = " ".join(map(str, raw_val)).lower()
                    if any(kw in joined for kw in specialty_keywords):
                        specialty_val = ", ".join(map(str, raw_val))
                    elif any(kw in joined for kw in category_keywords):
                        category_val = ", ".join(map(str, raw_val))
                    continue

                if isinstance(raw_val, dict):
                    continue

                str_val = str(raw_val)
                lower_val = str_val.lower()

                if '@' in str_val:
                    email_val = str_val
                    continue

                if lower_val in prefix_options:
                    prefix_val = str_val
                    continue

                if lower_val in food_options:
                    food_val = str_val
                    continue

                if any(kw in lower_val for kw in specialty_keywords) and len(lower_val) > 5:
                    specialty_val = str_val
                    continue
                if any(kw in lower_val for kw in category_keywords) and len(lower_val) > 5:
                    category_val = str_val
                    continue

                if re.match(r"^\+?[0-9\s\-]{9,15}$", str_val):
                    phone_val = str_val
                    continue

                if re.match(r"^\d{3,8}$", str_val):
                    reg_no_val = str_val
                    continue

                if any(kw in lower_val for kw in designation_keywords):
                    designation_val = str_val
                    continue

                if any(kw in lower_val for kw in council_keywords):
                    council_val = str_val
                    continue

            matched_strings = {email_val, phone_val, prefix_val, food_val, specialty_val, category_val, reg_no_val, designation_val, council_val}

            remaining_strings = []
            for val in values:
                if val is not None and val != '':
                    str_v = str(val)
                    if str_v not in matched_strings and '@' not in str_v and not re.match(r"^\+?[0-9\s\-]{9,15}$", str_v) and not isinstance(val, (list, dict)):
                        remaining_strings.append(str_v)

            for str_v in remaining_strings:
                lower = str_v.lower()
                if lower == (reg.participant_name or '').lower():
                    name_val = str_v
                elif any(kw in lower for kw in ['hospital', 'institute', 'clinic', 'college', 'university']) or lower in ['iqraa', 'hjkl', 'ygyg', 'sfds']:
                    institute_val = str_v
                elif not designation_val and any(kw in lower for kw in ['dr', 'doctor', 'consultant', 'physician', 'resident', 'student', 'intern', 'nurse', 'pharmacist', 'paramedic']):
                    designation_val = str_v
                elif not council_val and (any(kw in lower for kw in ['council', 'imc', 'mc']) or lower in ['hfgdg', 'jgfj', 'sf', 'gfhq1', 'hgfdh', 'sdfds']):
                    council_val = str_v
                else:
                    if not institute_val:
                        institute_val = str_v
                    elif not designation_val:
                        designation_val = str_v
                    elif not council_val:
                        council_val = str_v

            if any(kw in 'prefix' for kw in keywords): return prefix_val
            if any(kw in 'name' for kw in keywords): return name_val or reg.participant_name or ''
            if any(kw in 'email' for kw in keywords): return email_val or reg.participant_email or ''
            if any(kw in 'phone' for kw in keywords): return phone_val or reg.participant_phone or ''
            if any(kw in 'designation' for kw in keywords): return designation_val
            if any(kw in 'institute' for kw in keywords): return institute_val
            if any(kw in 'specialty' for kw in keywords): return specialty_val
            if any(kw in 'category' for kw in keywords): return category_val
            if any(kw in 'council' for kw in keywords): return council_val
            if any(kw in 'reg no' for kw in keywords): return reg_no_val
            if any(kw in 'food' for kw in keywords): return food_val

            return ''

        def get_val(keywords):
            # 1. Active form fields
            for field in form_fields:
                label_lower = field.label.lower()
                if any(kw in label_lower for kw in keywords):
                    val = (reg.field_data or {}).get(str(field.id))
                    if val is not None and val != '':
                        if isinstance(val, list):
                            return ", ".join(map(str, val))
                        elif isinstance(val, dict):
                            return val.get('name', val.get('url', str(val)))
                        return str(val)

            # 2. Historical fallback map
            field_data = reg.field_data or {}
            is_era2 = '31' in field_data
            is_era1 = '1' in field_data and '31' not in field_data and '65' not in field_data

            def format_val(val):
                if val is None:
                    return ''
                if isinstance(val, list):
                    return ", ".join(map(str, val))
                elif isinstance(val, dict):
                    return val.get('name', val.get('url', str(val)))
                return str(val)

            if is_era2:
                if any(kw in 'prefix' for kw in keywords): return format_val(field_data.get('30'))
                if any(kw in 'name' for kw in keywords): return format_val(field_data.get('31'))
                if any(kw in 'email' for kw in keywords): return format_val(field_data.get('32'))
                if any(kw in 'phone' for kw in keywords): return format_val(field_data.get('33'))
                if any(kw in 'designation' for kw in keywords): return format_val(field_data.get('34'))
                if any(kw in 'institute' for kw in keywords): return format_val(field_data.get('35'))
                if any(kw in 'specialty' for kw in keywords): return format_val(field_data.get('37'))
                if any(kw in 'category' for kw in keywords): return format_val(field_data.get('38'))
                if any(kw in 'council' for kw in keywords): return format_val(field_data.get('39') or field_data.get('36'))
                if any(kw in 'reg no' for kw in keywords): return format_val(field_data.get('41'))
                if any(kw in 'food' for kw in keywords): return format_val(field_data.get('40'))
            elif is_era1:
                if any(kw in 'name' for kw in keywords): return format_val(field_data.get('1'))
                if any(kw in 'email' for kw in keywords): return format_val(field_data.get('2'))
                if any(kw in 'phone' for kw in keywords): return format_val(field_data.get('3'))
                if any(kw in 'institute' for kw in keywords): return format_val(field_data.get('4'))
                if any(kw in 'designation' for kw in keywords): return format_val(field_data.get('5'))
                if any(kw in 'category' for kw in keywords): return format_val(field_data.get('6'))
                if any(kw in 'council' for kw in keywords): return format_val(field_data.get('7'))

            # 3. Dynamic Heuristic Fallback
            dynamic_val = get_val_dynamic(keywords)
            if dynamic_val:
                return dynamic_val

            # 4. Model level fallbacks
            if 'name' in keywords:
                return reg.participant_name or ''
            elif 'email' in keywords:
                return reg.participant_email or ''
            elif 'phone' in keywords:
                return reg.participant_phone or ''
            return ''

        row_data = [
            reg.registration_id or "PENDING",
            get_val(['name', 'fullname']),
            get_val(['email']),
            get_val(['phone', 'whatsapp']),
            get_val(['designation']),
            get_val(['institute', 'hospital']),
            get_val(['specialty', 'department']),
            get_val(['category']),
            get_val(['council']),
            get_val(['reg no', 'registration no']),
            get_val(['food']),
            amount,
            status,
            txnid,
            reg_date
        ]
        
        ws.append(row_data)
        
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = Font(name='Arial', size=10)
            cell.border = thin_border
            if col_idx in [1, 12, 13, 14, 15]:
                cell.alignment = Alignment(horizontal='center')
            else:
                cell.alignment = Alignment(horizontal='left')
        ws.row_dimensions[row_idx].height = 20
        
    auto_fit_columns(ws)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

def generate_articles_excel(articles):
    wb = Workbook()
    ws = wb.active
    ws.title = "Articles"
    
    headers = [
        "Article ID",
        "Registration ID",
        "Author Name",
        "Email",
        "Article Title",
        "Status",
        "Submitted Date",
        "Remarks",
        "Admin Remarks"
    ]
    style_header_row(ws, headers)
    
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    
    for row_idx, art in enumerate(articles, start=2):
        sub_date = art.submitted_date.astimezone(timezone.get_current_timezone()).strftime("%Y-%m-%d %H:%M")
        row_data = [
            art.id,
            art.registration.registration_id or "PENDING",
            art.author_name,
            art.email,
            art.article_title,
            art.status,
            sub_date,
            art.remarks or '',
            art.admin_remarks or ''
        ]
        ws.append(row_data)
        
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = Font(name='Arial', size=10)
            cell.border = thin_border
            if col_idx in [1, 2, 6, 7]:
                cell.alignment = Alignment(horizontal='center')
            else:
                cell.alignment = Alignment(horizontal='left')
        ws.row_dimensions[row_idx].height = 20
        
    auto_fit_columns(ws)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

def generate_revenue_excel(payments):
    wb = Workbook()
    ws = wb.active
    ws.title = "Revenue Tracking"
    
    headers = [
        "Transaction ID",
        "Registration ID",
        "Participant Name",
        "Amount",
        "Currency",
        "Status",
        "Payment Mode",
        "Payment Date"
    ]
    style_header_row(ws, headers)
    
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    
    for row_idx, pay in enumerate(payments, start=2):
        pay_date = pay.created_at.astimezone(timezone.get_current_timezone()).strftime("%Y-%m-%d %H:%M")
        row_data = [
            pay.transaction_id,
            pay.registration.registration_id or "PENDING",
            pay.registration.participant_name,
            pay.amount,
            pay.currency,
            pay.payment_status,
            pay.payment_mode or 'N/A',
            pay_date
        ]
        ws.append(row_data)
        
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = Font(name='Arial', size=10)
            cell.border = thin_border
            if col_idx in [1, 2, 4, 6, 8]:
                cell.alignment = Alignment(horizontal='center')
            else:
                cell.alignment = Alignment(horizontal='left')
        ws.row_dimensions[row_idx].height = 20
        
    auto_fit_columns(ws)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# --- PDF Generators (Landscape layout for wide reports) ---

def generate_registrations_pdf(registrations):
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=20,
        leftMargin=20,
        topMargin=20,
        bottomMargin=20
    )
    story = []
    styles = getSampleStyleSheet()
    
    from registration.models import RegistrationField, RegistrationForm
    active_form = RegistrationForm.objects.filter(is_active=True).first()
    form_fields = list(active_form.fields.all().order_by('order', 'id')) if active_form else []
    
    fs = 7
    ld = 9
    
    # Custom styles
    title_style = ParagraphStyle(
        'RepTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=18,
        textColor=colors.HexColor('#1a365d'),
        spaceAfter=15
    )
    cell_bold = ParagraphStyle(
        'RepCellBold',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=fs,
        leading=ld,
        textColor=colors.HexColor('#ffffff')
    )
    cell_normal = ParagraphStyle(
        'RepCellNormal',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=fs,
        leading=ld,
        textColor=colors.HexColor('#334155')
    )

    story.append(Paragraph("TOXIQ Program - Registration Report", title_style))
    story.append(Paragraph(f"Generated on: {timezone.now().strftime('%d-%b-%Y %H:%M')}", cell_normal))
    story.append(Spacer(1, 10))

    headers = [
        Paragraph("Reg ID", cell_bold),
        Paragraph("Full Name", cell_bold),
        Paragraph("Email", cell_bold),
        Paragraph("Phone", cell_bold),
        Paragraph("Designation", cell_bold),
        Paragraph("Institute", cell_bold),
        Paragraph("Specialty", cell_bold),
        Paragraph("Category", cell_bold),
        Paragraph("Council", cell_bold),
        Paragraph("Reg No", cell_bold),
        Paragraph("Food", cell_bold),
        Paragraph("Amount", cell_bold),
        Paragraph("Status", cell_bold),
        Paragraph("Date", cell_bold)
    ]
    
    table_data = [headers]
    for reg in registrations:
        payment = reg.payments.filter(payment_status='SUCCESS').first() or reg.payments.first()
        status = payment.payment_status if payment else "N/A"
        amount = f"{payment.currency or 'INR'} {float(payment.amount):.2f}" if payment and payment.amount else "N/A"
        reg_date = reg.created_at.astimezone(timezone.get_current_timezone()).strftime("%Y-%m-%d")
        
        def get_val_dynamic(keywords):
            field_data = reg.field_data or {}
            values = list(field_data.values())

            prefix_options = ['mr.', 'ms.', 'mrs.', 'dr.', 'mr', 'ms', 'mrs', 'dr']
            food_options = ['veg', 'non-veg', 'veg.', 'non-veg.']
            specialty_keywords = ['emergency', 'pharmacy', 'critical', 'care', 'medicine', 'pediatrics', 'forensic', 'practitioner', 'specialty', 'department']
            category_keywords = ['speaker', 'consultant', 'resident', 'student', 'intern', 'nurse', 'pharmacist', 'paramedic', 'delegate', 'category']
            designation_keywords = ['doctor', 'consultant', 'physician', 'resident', 'student', 'intern', 'nurse', 'pharmacist', 'paramedic', 'professor', 'lecturer', 'director', 'manager', 'specialist', 'practitioner', 'hod', 'registrar', 'fellow', 'designation']
            council_keywords = ['council', 'medical', 'imc', 'kcl', 'mc', 'karnataka', 'delhi', 'maharashtra', 'registration council', 'state']

            email_val = ''
            phone_val = ''
            prefix_val = ''
            food_val = ''
            specialty_val = ''
            category_val = ''
            reg_no_val = ''
            designation_val = ''
            council_val = ''
            name_val = ''
            institute_val = ''

            import re

            for raw_val in values:
                if raw_val is None or raw_val == '':
                    continue

                if isinstance(raw_val, list):
                    joined = " ".join(map(str, raw_val)).lower()
                    if any(kw in joined for kw in specialty_keywords):
                        specialty_val = ", ".join(map(str, raw_val))
                    elif any(kw in joined for kw in category_keywords):
                        category_val = ", ".join(map(str, raw_val))
                    continue

                if isinstance(raw_val, dict):
                    continue

                str_val = str(raw_val)
                lower_val = str_val.lower()

                if '@' in str_val:
                    email_val = str_val
                    continue

                if lower_val in prefix_options:
                    prefix_val = str_val
                    continue

                if lower_val in food_options:
                    food_val = str_val
                    continue

                if any(kw in lower_val for kw in specialty_keywords) and len(lower_val) > 5:
                    specialty_val = str_val
                    continue
                if any(kw in lower_val for kw in category_keywords) and len(lower_val) > 5:
                    category_val = str_val
                    continue

                if re.match(r"^\+?[0-9\s\-]{9,15}$", str_val):
                    phone_val = str_val
                    continue

                if re.match(r"^\d{3,8}$", str_val):
                    reg_no_val = str_val
                    continue

                if any(kw in lower_val for kw in designation_keywords):
                    designation_val = str_val
                    continue

                if any(kw in lower_val for kw in council_keywords):
                    council_val = str_val
                    continue

            matched_strings = {email_val, phone_val, prefix_val, food_val, specialty_val, category_val, reg_no_val, designation_val, council_val}

            remaining_strings = []
            for val in values:
                if val is not None and val != '':
                    str_v = str(val)
                    if str_v not in matched_strings and '@' not in str_v and not re.match(r"^\+?[0-9\s\-]{9,15}$", str_v) and not isinstance(val, (list, dict)):
                        remaining_strings.append(str_v)

            for str_v in remaining_strings:
                lower = str_v.lower()
                if lower == (reg.participant_name or '').lower():
                    name_val = str_v
                elif any(kw in lower for kw in ['hospital', 'institute', 'clinic', 'college', 'university']) or lower in ['iqraa', 'hjkl', 'ygyg', 'sfds']:
                    institute_val = str_v
                elif not designation_val and any(kw in lower for kw in ['dr', 'doctor', 'consultant', 'physician', 'resident', 'student', 'intern', 'nurse', 'pharmacist', 'paramedic']):
                    designation_val = str_v
                elif not council_val and (any(kw in lower for kw in ['council', 'imc', 'mc']) or lower in ['hfgdg', 'jgfj', 'sf', 'gfhq1', 'hgfdh', 'sdfds']):
                    council_val = str_v
                else:
                    if not institute_val:
                        institute_val = str_v
                    elif not designation_val:
                        designation_val = str_v
                    elif not council_val:
                        council_val = str_v

            if any(kw in 'prefix' for kw in keywords): return prefix_val
            if any(kw in 'name' for kw in keywords): return name_val or reg.participant_name or ''
            if any(kw in 'email' for kw in keywords): return email_val or reg.participant_email or ''
            if any(kw in 'phone' for kw in keywords): return phone_val or reg.participant_phone or ''
            if any(kw in 'designation' for kw in keywords): return designation_val
            if any(kw in 'institute' for kw in keywords): return institute_val
            if any(kw in 'specialty' for kw in keywords): return specialty_val
            if any(kw in 'category' for kw in keywords): return category_val
            if any(kw in 'council' for kw in keywords): return council_val
            if any(kw in 'reg no' for kw in keywords): return reg_no_val
            if any(kw in 'food' for kw in keywords): return food_val

            return ''

        def get_val(keywords):
            # 1. Active form fields
            for field in form_fields:
                label_lower = field.label.lower()
                if any(kw in label_lower for kw in keywords):
                    val = (reg.field_data or {}).get(str(field.id))
                    if val is not None and val != '':
                        if isinstance(val, list):
                            return ", ".join(map(str, val))
                        elif isinstance(val, dict):
                            return val.get('name', val.get('url', str(val)))
                        return str(val)

            # 2. Historical fallback map
            field_data = reg.field_data or {}
            is_era2 = '31' in field_data
            is_era1 = '1' in field_data and '31' not in field_data and '65' not in field_data

            def format_val(val):
                if val is None:
                    return ''
                if isinstance(val, list):
                    return ", ".join(map(str, val))
                elif isinstance(val, dict):
                    return val.get('name', val.get('url', str(val)))
                return str(val)

            if is_era2:
                if any(kw in 'prefix' for kw in keywords): return format_val(field_data.get('30'))
                if any(kw in 'name' for kw in keywords): return format_val(field_data.get('31'))
                if any(kw in 'email' for kw in keywords): return format_val(field_data.get('32'))
                if any(kw in 'phone' for kw in keywords): return format_val(field_data.get('33'))
                if any(kw in 'designation' for kw in keywords): return format_val(field_data.get('34'))
                if any(kw in 'institute' for kw in keywords): return format_val(field_data.get('35'))
                if any(kw in 'specialty' for kw in keywords): return format_val(field_data.get('37'))
                if any(kw in 'category' for kw in keywords): return format_val(field_data.get('38'))
                if any(kw in 'council' for kw in keywords): return format_val(field_data.get('39') or field_data.get('36'))
                if any(kw in 'reg no' for kw in keywords): return format_val(field_data.get('41'))
                if any(kw in 'food' for kw in keywords): return format_val(field_data.get('40'))
            elif is_era1:
                if any(kw in 'name' for kw in keywords): return format_val(field_data.get('1'))
                if any(kw in 'email' for kw in keywords): return format_val(field_data.get('2'))
                if any(kw in 'phone' for kw in keywords): return format_val(field_data.get('3'))
                if any(kw in 'institute' for kw in keywords): return format_val(field_data.get('4'))
                if any(kw in 'designation' for kw in keywords): return format_val(field_data.get('5'))
                if any(kw in 'category' for kw in keywords): return format_val(field_data.get('6'))
                if any(kw in 'council' for kw in keywords): return format_val(field_data.get('7'))

            # 3. Dynamic Heuristic Fallback
            dynamic_val = get_val_dynamic(keywords)
            if dynamic_val:
                return dynamic_val

            # 4. Model level fallbacks
            if 'name' in keywords:
                return reg.participant_name or ''
            elif 'email' in keywords:
                return reg.participant_email or ''
            elif 'phone' in keywords:
                return reg.participant_phone or ''
            return ''

        row = [
            Paragraph(reg.registration_id or "PENDING", cell_normal),
            Paragraph(get_val(['name', 'fullname']), cell_normal),
            Paragraph(get_val(['email']), cell_normal),
            Paragraph(get_val(['phone', 'whatsapp']), cell_normal),
            Paragraph(get_val(['designation']), cell_normal),
            Paragraph(get_val(['institute', 'hospital']), cell_normal),
            Paragraph(get_val(['specialty', 'department']), cell_normal),
            Paragraph(get_val(['category']), cell_normal),
            Paragraph(get_val(['council']), cell_normal),
            Paragraph(get_val(['reg no', 'registration no']), cell_normal),
            Paragraph(get_val(['food']), cell_normal),
            Paragraph(amount, cell_normal),
            Paragraph(status, cell_normal),
            Paragraph(reg_date, cell_normal)
        ]
        table_data.append(row)
        
    col_widths = [50, 80, 85, 60, 60, 65, 65, 65, 50, 45, 35, 45, 45, 50]
        
    t = Table(table_data, colWidths=col_widths)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1A365D')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#94a3b8')),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        ('PADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(t)
    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()

def generate_articles_pdf(articles):
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30
    )
    story = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'RepTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=18,
        textColor=colors.HexColor('#1a365d'),
        spaceAfter=15
    )
    cell_bold = ParagraphStyle(
        'RepCellBold',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        leading=11,
        textColor=colors.HexColor('#ffffff')
    )
    cell_normal = ParagraphStyle(
        'RepCellNormal',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        leading=10,
        textColor=colors.HexColor('#334155')
    )

    story.append(Paragraph("TOXIQ Program - Article Submissions Moderation Report", title_style))
    story.append(Paragraph(f"Generated on: {timezone.now().strftime('%d-%b-%Y %H:%M')}", cell_normal))
    story.append(Spacer(1, 10))

    headers = [
        Paragraph("ID", cell_bold),
        Paragraph("Reg ID", cell_bold),
        Paragraph("Author Name", cell_bold),
        Paragraph("Email", cell_bold),
        Paragraph("Article Title", cell_bold),
        Paragraph("Status", cell_bold),
        Paragraph("Date", cell_bold)
    ]
    
    table_data = [headers]
    for art in articles:
        sub_date = art.submitted_date.astimezone(timezone.get_current_timezone()).strftime("%Y-%m-%d")
        table_data.append([
            Paragraph(str(art.id), cell_normal),
            Paragraph(art.registration.registration_id or "PENDING", cell_normal),
            Paragraph(art.author_name[:25], cell_normal),
            Paragraph(art.email[:25], cell_normal),
            Paragraph(art.article_title[:45], cell_normal),
            Paragraph(art.status, cell_normal),
            Paragraph(sub_date, cell_normal),
        ])
        
    t = Table(table_data, colWidths=[35, 75, 120, 130, 200, 75, 65])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1A365D')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#94a3b8')),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(t)
    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()

def generate_revenue_pdf(payments):
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30
    )
    story = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'RepTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=18,
        textColor=colors.HexColor('#1a365d'),
        spaceAfter=15
    )
    cell_bold = ParagraphStyle(
        'RepCellBold',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        leading=11,
        textColor=colors.HexColor('#ffffff')
    )
    cell_normal = ParagraphStyle(
        'RepCellNormal',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        leading=10,
        textColor=colors.HexColor('#334155')
    )

    story.append(Paragraph("TOXIQ Program - Revenue Tracking Report", title_style))
    story.append(Paragraph(f"Generated on: {timezone.now().strftime('%d-%b-%Y %H:%M')}", cell_normal))
    story.append(Spacer(1, 10))

    headers = [
        Paragraph("Transaction ID", cell_bold),
        Paragraph("Reg ID", cell_bold),
        Paragraph("Name", cell_bold),
        Paragraph("Amount", cell_bold),
        Paragraph("Status", cell_bold),
        Paragraph("Mode", cell_bold),
        Paragraph("Date", cell_bold)
    ]
    
    table_data = [headers]
    for pay in payments:
        pay_date = pay.created_at.astimezone(timezone.get_current_timezone()).strftime("%Y-%m-%d %H:%M")
        table_data.append([
            Paragraph(pay.transaction_id, cell_normal),
            Paragraph(pay.registration.registration_id or "PENDING", cell_normal),
            Paragraph(pay.registration.participant_name[:30], cell_normal),
            Paragraph(f"{pay.currency} {pay.amount:.2f}", cell_normal),
            Paragraph(pay.payment_status, cell_normal),
            Paragraph(pay.payment_mode or 'N/A', cell_normal),
            Paragraph(pay_date, cell_normal),
        ])
        
    t = Table(table_data, colWidths=[110, 80, 160, 80, 80, 85, 110])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1A365D')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#94a3b8')),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(t)
    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()
